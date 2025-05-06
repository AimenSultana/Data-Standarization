import boto3
import pandas as pd
from datetime import datetime
import os
from typing import List, Dict
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class S3BucketAnalyzer:
    def __init__(self, aws_access_key_id: str, aws_secret_access_key: str):
        self.s3 = boto3.client(
            's3',
            aws_access_key_id=aws_access_key_id,
            aws_secret_access_key=aws_secret_access_key
        )
    
    def get_folder_structure(self, bucket_name: str, prefix: str = '') -> Dict:
        """Get detailed folder structure with metadata"""
        structure = {}
        paginator = self.s3.get_paginator('list_objects_v2')
        
        try:
            for page in paginator.paginate(Bucket=bucket_name, Prefix=prefix, Delimiter='/'):
                # Process folders
                for folder in page.get('CommonPrefixes', []):
                    folder_path = folder['Prefix']
                    folder_name = folder_path.rstrip('/').split('/')[-1]
                    structure[folder_path] = {
                        'Type': 'Folder',
                        'Name': folder_name,
                        'Path': folder_path,
                        'Contents': self.get_folder_structure(bucket_name, folder_path)
                    }
                
                # Process files
                for obj in page.get('Contents', []):
                    if not obj['Key'].endswith('/'):  # Skip folder markers
                        file_path = obj['Key']
                        file_name = file_path.split('/')[-1]
                        file_ext = os.path.splitext(file_name)[1].lower()[1:] or 'No Extension'
                        
                        structure[file_path] = {
                            'Type': 'File',
                            'Name': file_name,
                            'Path': file_path,
                            'Extension': file_ext,
                            'Size': obj['Size'],
                            'Last Modified': obj['LastModified'].replace(tzinfo=None),
                            'Storage Class': obj.get('StorageClass', 'STANDARD')
                        }
        except Exception as e:
            print(f"Error accessing bucket: {e}")
            return {}

        return structure

    def flatten_structure(self, structure: Dict, parent_path: str = '') -> List[Dict]:
        """Flatten the nested folder structure for Excel output"""
        flat_data = []
        
        for path, item in structure.items():
            if item['Type'] == 'Folder':
                # Count files and subfolders
                file_count = 0
                folder_count = 0
                for content in item['Contents'].values():
                    if content['Type'] == 'File':
                        file_count += 1
                    elif content['Type'] == 'Folder':
                        folder_count += 1
                
                # Add folder entry
                flat_data.append({
                    'Level': path.count('/') - 1,
                    'Type': 'Folder',
                    'Name': item['Name'],
                    'Full Path': path,
                    'Size': '',
                    'File Count': file_count,
                    'Folder Count': folder_count,
                    'Last Modified': '',
                    'Storage Class': ''
                })
                # Add contents recursively
                flat_data.extend(self.flatten_structure(item['Contents'], path))
            else:
                # Add file entry
                flat_data.append({
                    'Level': path.count('/'),
                    'Type': 'File',
                    'Name': item['Name'],
                    'Full Path': path,
                    'Size': item['Size'],
                    'File Count': '',
                    'Folder Count': '',
                    'Last Modified': item['Last Modified'],
                    'Storage Class': item['Storage Class'],
                    'Extension': item['Extension']
                })
        
        return flat_data

    def generate_report(self, bucket_name: str, output_path: str = None, specific_folder: str = None):
        """Generate comprehensive Excel report"""
        print(f"\nAnalyzing {'folder' if specific_folder else 'bucket'}...")
        
        # Get the folder structure
        prefix = specific_folder if specific_folder else ''
        structure = self.get_folder_structure(bucket_name, prefix)
        
        if not structure:
            print("No files/folders found in the specified location.")
            return
        
        # Flatten structure for Excel
        flat_data = self.flatten_structure(structure)
        df = pd.DataFrame(flat_data)
        
        if df.empty:
            print("No data to generate report.")
            return
        
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_part = f"_{specific_folder.replace('/', '_').rstrip('_')}" if specific_folder else '_full_bucket'
        output_filename = output_path or f"{bucket_name}{folder_part}_inventory_{timestamp}.xlsx"
        
        # Create Excel writer
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            # Create the main structure sheet
            self._create_structure_sheet(writer, df, bucket_name, prefix)
            
            # Create summary sheets
            self._create_summary_sheets(writer, df)
            
            # Create folder tree sheet
            self._create_folder_tree_sheet(writer, structure)
        
        print(f"\nReport generated successfully: {output_filename}")
        print(f"Total items analyzed: {len(df)}")
        print(f"Folders: {len(df[df['Type'] == 'Folder'])}")
        print(f"Files: {len(df[df['Type'] == 'File'])}")

    def _format_size(self, size_bytes):
        """Convert bytes to human-readable format like S3"""
        if size_bytes == 0:
            return '0 KB'
        for unit in ['Bytes', 'KB', 'MB', 'GB', 'TB']:
            if size_bytes < 1024.0:
                if unit == 'Bytes':
                    return f"{int(size_bytes)} {unit}"
                else:
                    return f"{size_bytes:.2f} {unit}"
            size_bytes /= 1024.0
        return f"{size_bytes:.2f} PB"

    def _create_structure_sheet(self, writer, df, bucket_name, prefix):
        """Create the main structure sheet with enhanced formatting"""
        # Prepare data for display
        display_data = []
        for _, row in df.iterrows():
            indent = '    ' * row['Level']
            display_name = f"{indent}{row['Name']}{'/' if row['Type'] == 'Folder' else ''}"
            
            last_modified = ''
            if row['Type'] == 'File' and pd.notna(row['Last Modified']):
                last_modified = row['Last Modified'].strftime('%Y-%m-%d %H:%M') if not isinstance(row['Last Modified'], str) else row['Last Modified']
            
            is_empty = (row['Type'] == 'Folder' and row['File Count'] == 0 and row['Folder Count'] == 0)
            
            display_data.append({
                'Structure': display_name,
                'Format': row['Extension'].upper() if row['Type'] == 'File' else '',
                'Type': row['Type'],
                'Size': self._format_size(row['Size']) if row['Type'] == 'File' else '',
                'Last Modified': last_modified,
                'File Count': row['File Count'] if row['Type'] == 'Folder' else '',
                'Folder Count': row['Folder Count'] if row['Type'] == 'Folder' else '',
                '_IsEmpty': is_empty  # Hidden flag for formatting
            })
        
        display_df = pd.DataFrame(display_data)
        # Reorder columns to put Format before Type
        display_df = display_df[['Structure', 'Format', 'Type', 'Size', 'Last Modified', 'File Count', 'Folder Count']]
        display_df.to_excel(writer, sheet_name='Structure', index=False)
        
        # Apply enhanced formatting
        workbook = writer.book
        worksheet = writer.sheets['Structure']
        
        # Set column widths (updated order)
        column_widths = {
            'A': 40,  # Structure
            'B': 10,  # Format
            'C': 10,  # Type
            'D': 15,  # Size (dynamic based on content)
            'E': 20,  # Last Modified
            'F': 12,  # File Count
            'G': 12   # Folder Count
        }
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Define styles
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        folder_font = Font(bold=True)
        empty_folder_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Format header
        for col in range(1, len(column_widths) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Format data rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), 2):
            is_empty = display_data[row_idx-2]['_IsEmpty']
            is_folder = worksheet.cell(row=row_idx, column=3).value == 'Folder'  # Type column is now C
            
            for cell in row:
                cell.border = border
                
                # Right-align size values
                if cell.column_letter == 'D':  # Size column
                    cell.alignment = Alignment(horizontal='right')
                # Apply special formatting for folders
                elif cell.column_letter == 'A':  # Structure column
                    if is_folder:
                        cell.font = folder_font
                    cell.alignment = Alignment(horizontal='left', indent=cell.value.count('    '))
                elif cell.column_letter == 'B':  # Format column
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='center')
                
                # Highlight empty folders in red (Structure and Type columns)
                if is_empty and cell.column_letter in ['A','B','C']:
                    cell.fill = empty_folder_fill
        
        # Auto-adjust size column width based on content
        max_length = max(display_df['Size'].astype(str).apply(len).max(), len('Size'))
        worksheet.column_dimensions['D'].width = max_length + 2
        
        # Add title
        title = f"S3 Bucket Structure: {bucket_name}"
        if prefix:
            title += f" (Folder: {prefix})"
        worksheet.insert_rows(1)
        worksheet.merge_cells('A1:G1')  # A to G for all columns
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.border = border

    def _create_summary_sheets(self, writer, df):
        """Create summary sheets"""
        # File type summary
        if not df[df['Type'] == 'File'].empty:
            file_summary = df[df['Type'] == 'File'].groupby('Extension', observed=False).agg({
                'Name': 'count',
                'Size': 'sum'
            }).rename(columns={'Name': 'File Count', 'Size': 'Total Size (Bytes)'})
            file_summary['Total Size (MB)'] = file_summary['Total Size (Bytes)'] / (1024 * 1024)
            file_summary.to_excel(writer, sheet_name='File Type Summary')
        
        # Folder summary
        if not df[df['Type'] == 'Folder'].empty:
            folder_summary = df[df['Type'] == 'Folder'].copy()
            folder_summary['Depth'] = folder_summary['Full Path'].str.count('/')
            folder_summary.to_excel(writer, sheet_name='Folder Summary', index=False)
        
        # Size distribution
        if not df[df['Type'] == 'File'].empty:
            size_bins = [0, 1024, 10240, 102400, 1048576, 10485760, 104857600, float('inf')]
            size_labels = ['<1KB', '1-10KB', '10-100KB', '100KB-1MB', '1-10MB', '10-100MB', '>100MB']
            size_dist = df[df['Type'] == 'File'].copy()
            size_dist['Size Range'] = pd.cut(size_dist['Size'], bins=size_bins, labels=size_labels)
            size_dist = size_dist.groupby('Size Range', observed=False).agg({'Name': 'count'}).rename(
                columns={'Name': 'File Count'})
            size_dist['Percentage'] = (size_dist['File Count'] / size_dist['File Count'].sum()) * 100
            size_dist.to_excel(writer, sheet_name='Size Distribution')

    def _create_folder_tree_sheet(self, writer, structure, max_depth=10):
        """Create folder tree view"""
        tree_data = []
        self._build_tree_data(structure, tree_data, max_depth)
        
        if tree_data:
            tree_df = pd.DataFrame(tree_data)
            tree_df.to_excel(writer, sheet_name='Folder Tree', index=False)

    def _build_tree_data(self, structure, tree_data, max_depth, current_depth=0, path=''):
        """Build tree data recursively"""
        if current_depth > max_depth:
            return
            
        for item_path, item in structure.items():
            if item['Type'] == 'Folder':
                indent = '    ' * current_depth
                tree_data.append({
                    'Level': current_depth,
                    'Tree View': f"{indent}📁 {item['Name']}",
                    'File Count': item.get('File Count', ''),
                    'Folder Count': item.get('Folder Count', '')
                })
                self._build_tree_data(
                    item['Contents'], 
                    tree_data, 
                    max_depth, 
                    current_depth + 1, 
                    item_path
                )

def main():
    print("S3 Bucket Analyzer\n")
    print("This tool will generate a detailed inventory report of your S3 bucket or a specific project folder.\n")
    
    # Get AWS credentials
    aws_access_key_id = input("Enter your AWS Access Key ID: ").strip()
    aws_secret_access_key = input("Enter your AWS Secret Access Key: ").strip()
    
    # Get bucket name
    bucket_name = input("Enter the S3 bucket name: ").strip()
    
    # Ask if user wants to analyze specific folder
    while True:
        choice = input("Do you want to analyze:\n1. The entire bucket\n2. A specific project folder\nEnter 1 or 2: ").strip()
        if choice in ['1', '2']:
            break
        print("Invalid choice. Please enter 1 or 2")
    
    specific_folder = ''
    if choice == '2':
        specific_folder = input("\nEnter the project folder name (e.g., 'TRR' or 'projects/TRR/'): ").strip()
        if not specific_folder.endswith('/'):
            specific_folder += '/'
        print(f"\nWill analyze only the folder: {specific_folder}")
    else:
        print("\nWill analyze the entire bucket")
    
    # Initialize analyzer
    analyzer = S3BucketAnalyzer(aws_access_key_id, aws_secret_access_key)
    
    # Generate report
    analyzer.generate_report(bucket_name, specific_folder=specific_folder)

if __name__ == "__main__":
    main()